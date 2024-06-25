
# 检查统计废弃表的资源ID在所有项目中使用了多少次，输出到 【废弃资源使用统计.xlsx】中

import os
import time
import chardet
import logging

import openpyxl

# 用于存放废弃资源的GUID
DiscardedGuidArray = []

# 每个游戏工程项目需要检查的文件夹
ChangePath = ["JavaScripts", "Prefabs",
              "TempPrefabs", "Levels", "UI", "Materials"]

XlsxFilePath = 'dist\废弃表.xlsx'

# 用于存放项目路径
projects = []

all_path = []

resList = {}

tsList = {}

pnum = int(0)


# 读取废弃表，将所有的GUID存入DiscardedGuidArray
def xlsx_init():
    if not os.path.exists(XlsxFilePath):
        return False
    wb = openpyxl.load_workbook(XlsxFilePath, data_only=True)
    sheetNames = wb.sheetnames
    for sheetname in sheetNames:
        sheet = wb[sheetname]
        rowCount = sheet.max_row
        # 从第二行开始读取
        for row in range(2, rowCount + 1):
            cellValue = sheet.cell(row, 1).value
            if cellValue == None:
                continue
            guid = str(cellValue)
            if guid in DiscardedGuidArray:
                continue
            DiscardedGuidArray.append(guid)
            continue
    print('废弃资源数量:' + str(DiscardedGuidArray.__len__()))
    return True


# 获取文件编码
def get_file_encoding(in_path):
    f = open(in_path, 'rb')
    data = f.read()
    file_encoding = chardet.detect(data).get('encoding')
    f.close()
    return file_encoding

# 获取所有文件路径
def get_all_file_path(in_path):
    if not os.path.exists(in_path):
        return

    path_list = os.listdir(in_path)
    for p in path_list:
        fullpath = os.path.join(in_path, p)
        if os.path.isdir(fullpath):
            get_all_file_path(fullpath)
            continue
        if fullpath.endswith('.level') or fullpath.endswith('.mat') or fullpath.endswith('.ui') or fullpath.endswith('.prefab') or fullpath.endswith('.ts'):
            all_path.append(fullpath)

# 初始化项目路径
def init_path(project):
    global pnum
    pnum += 1
    print(str(pnum)+'项目路径:' + project)
    for p in ChangePath:
        get_all_file_path(os.path.join(project, p))

# 检查文件中是否包含废弃资源的GUID
def scan_file(in_path: str):
    if not os.path.exists(in_path):
        return
    if in_path.endswith('.d.ts') or in_path.endswith('.meta'):
        return
    encode = get_file_encoding(in_path)
    if encode == 'Windows-1254':
        encode = 'utf-8'
    print("检测文件：", in_path, str(encode))  # 修改了这里，原来的变量名path可能是错误的，应该是in_path
    try:
        with open(in_path, "r", encoding=encode, errors='ignore') as f:  # 添加了errors='ignore来忽略无法解码的字符
            content = f.read()
    except UnicodeDecodeError:
        print(f"Could not read file {in_path} using encoding {encode}. Skipping this file.")
        return
    for g in DiscardedGuidArray:
        gstr = '\"' + g + '\"'
        tc = content.find(gstr)
        if in_path.endswith('.ts') and tc < 0:
            tc = content.find('\"' + g + '\\')
        if tc != -1:
            key = str(g)
            num = 0
            if key in resList.keys():
                num = resList[key]
            count = content.count(gstr)
            num += count
            resList[key] = num
            if in_path.endswith('.ts'):
                tsn = count
                if key in tsList.keys():
                    tsn += tsList[key]
                tsList[key] = tsn

# 初始化项目路径
def init_file(ps):
    list = os.listdir(ps)
    for l in list:
        full_path = os.path.join(ps, l)  # 形成完整的路径
        if os.path.isdir(full_path):  # 检查完整的路径是否为目录
            projects.append(l)


if __name__ == '__main__':
    # 创建 logger 对象
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 创建 console handler 并设置级别为 DEBUG
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)

    a = time.time()
    if xlsx_init():
        root = r"C:\Users\admin\Downloads\online\games\projects1"
        xlsx = 'dist\废弃资源使用统计.xlsx'
        book = openpyxl.load_workbook(xlsx, data_only=True)
        sheet = book['list']
        col = ['资源GUID', '使用次数', 'ts中的次数']
        book.remove(sheet)
        sheet = book.create_sheet('list', 0)
        sheet.append(col)

        init_file(root)
        for p in projects:
            file = str(p)
            all_path = []
            init_path(root + '\\'+file)
            for path in all_path:
                scan_file(path)
        print('废弃资源使用数量:' + str(resList.items().__len__()))
        for k, v in resList.items():
            ts = 0
            if k in tsList.keys():
                ts = tsList[k]
            sheet.append([k, v, ts])
        book.save(xlsx)
    else:
        print('文件:[' + XlsxFilePath + ']不存在')
    time.sleep(1)
    os.system('pause')
