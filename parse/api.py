import os
import time
import chardet
import openpyxl

apiList = ['setEyeHighLightTexture', 'setEyeHighLightColor']

files = ["JavaScripts", "Prefabs"]

projects = []
all_path = []
useCount = []


def get_file_encoding(in_path):
    f = open(in_path, 'rb')
    data = f.read()
    file_encoding = chardet.detect(data).get('encoding')
    f.close()
    return file_encoding


def get_all_file_path(in_path):
    if not os.path.exists(in_path):
        return

    path_list = os.listdir(in_path)
    for p in path_list:
        fullpath = os.path.join(in_path, p)
        if os.path.isdir(fullpath):
            get_all_file_path(fullpath)
            continue
        if fullpath.endswith('.ts'):
            all_path.append(fullpath)


def init_path(project):
    print('项目路径:' + project)
    for p in files:
        get_all_file_path(os.path.join(project, p))
    if all_path.__len__() < 1:
        print('空项目:' + project)


def scan_file(in_path: str):
    if not os.path.exists(in_path):
        return
    if in_path.endswith('.d.ts'):
        return
    encode = get_file_encoding(in_path)
    if encode == 'Windows-1254':
        encode = 'utf-8'

    # print(path + ' encoding :' + str(encode))
    with open(in_path, "r", encoding=encode) as f:
        content = f.read()
        # print(content)
        idx = 0
        for g in apiList:
            num = content.count(g)
            useCount[idx] += num
            idx += 1


def init_file(ps):
    list = os.listdir(ps)
    for l in list:
        l = os.path.join(ps, l)
        if os.path.isdir(l):
            projects.append(l)


if __name__ == '__main__':
    st = time.time()
    root = r"C:\Users\admin\Downloads\online\games\projects1"
    print(root, ' API检测中...')
    xlsx = 'dist\项目API使用统计.xlsx'

    if not os.path.exists(xlsx):
        book = openpyxl.Workbook()
        book.save(xlsx)
    else:
        book = openpyxl.load_workbook(xlsx, data_only=True)

    if 'list' in book.sheetnames:
        sheet = book['list']
    else:
        sheet = book.create_sheet('list')
    col = ["序号", '项目名']
    for api in apiList:
        col.append(api)
    book.remove(sheet)
    sheet = book.create_sheet('list', 0)
    sheet.append(col)
    init_file(root)
    idx = 0
    for p in projects:
        print('项目:', p)
        idx += 1
        file = str(p)
        all_path = []
        useCount = []
        for api in apiList:
            useCount.append(0)
        book.save(xlsx)
        if file.find('.') < 0:
            init_path(root + '\\'+file)
            for path in all_path:
                scan_file(path)
            useCount.insert(0, file)
            id = idx
            if all_path.__len__() < 1:
                idx -= 1
                id = 0
            useCount.insert(0, id)
            sheet.append(useCount)
    ut = time.time()-st
    print('API检测完成,用时:', ut)
