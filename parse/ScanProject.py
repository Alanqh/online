import os
import time
import chardet
import openpyxl

checkFiles = ["JavaScripts", "Prefabs", "TempPrefabs", "Levels", "UI", "Materials"]
delList = []
delFile = 'dist/废弃表.xlsx'
projects = []

all_path = []
nums = []
resList = []


def xlsx_init():
    if not os.path.exists(delFile):
        return False
    wb = openpyxl.load_workbook(delFile, data_only=True)
    sheetNames = wb.sheetnames
    for sheetname in sheetNames:
        sheet = wb[sheetname]
        rowCount = sheet.max_row
        for row in range(2, rowCount + 1):
            cellValue = sheet.cell(row, 1).value
            if cellValue == None:
                continue
            delList.append(str(cellValue))
            continue
    return True


def get_file_encoding(in_path):
    encod = 'utf-8'
    with open(in_path, 'rb') as f:
        coding = chardet.detect(f.read())
        encod = coding['encoding']
        f.close()
    return encod


def get_all_file_path(in_path):
    if not os.path.exists(in_path):
        return

    path_list = os.listdir(in_path)
    for p in path_list:
        fullpath = os.path.join(in_path, p)
        if os.path.isdir(fullpath):
            get_all_file_path(fullpath)
            continue
        if fullpath.endswith('.level') or fullpath.endswith('.mat') or fullpath.endswith('.ui') or fullpath.endswith(
                '.prefab') or fullpath.endswith('.ts'):
            all_path.append(fullpath)


def init_path(project):
    print('检查项目:' + project)
    for p in checkFiles:
        get_all_file_path(os.path.join(project, p))


def scan_file(in_path: str):
    if not os.path.exists(in_path):
        return
    if in_path.endswith('.d.ts'):
        return
    encode = get_file_encoding(in_path)
    if encode == 'Windows-1254':
        encode = 'utf-8'

    try:
        with open(in_path, "r", encoding=encode, errors='ignore') as f:
            content = f.read()
            for g in delList:
                gstr = '\"' + g + '\"'
                num1 = 0
                num2 = 0
                if content.find(gstr) != -1:
                    num1 = content.count(gstr)
                if in_path.endswith('.level'):
                    nums[0] = nums[0] + num1 + num2
                elif in_path.endswith('.prefab'):
                    nums[1] = nums[1] + num1 + num2
                elif in_path.endswith('.mat'):
                    nums[2] = nums[2] + num1 + num2
                elif in_path.endswith('.ui'):
                    nums[3] = nums[3] + num1 + num2
                elif in_path.endswith('.ts'):
                    nums[4] = nums[4] + num1 + num2
                # else:
                #     nums[5]=nums[5]+num1+num2
                if (num1 + num2 > 0) and (g not in resList):
                    resList.append(g)

    except UnicodeDecodeError:
        print(f"Could not read file {in_path} using encoding {encode}. Skipping this file.")



def init_file(ps):
    list = os.listdir(ps)
    for l in list:
        full_path = os.path.join(ps, l)  # 形成完整的路径
        if os.path.isdir(full_path):  # 检查完整的路径是否为目录
            projects.append(l)




if __name__ == '__main__':

    a = time.time()
    if xlsx_init():
        # root = os.getcwd()
        # print(root)
        root = r"C:\Users\admin\Downloads\online\games\projects1"

        xlsx = 'dist\项目废弃资源使用统计.xlsx'
        book = openpyxl.load_workbook(xlsx, data_only=True)
        sheet = book['list']
        col = ['项目名', 'level文件', 'prefab文件',
               '材质文件', 'UI文件', 'TS文件', '总计', '废弃资源数']
        book.remove(sheet)
        sheet = book.create_sheet('list', 0)
        sheet.append(col)

        init_file(root)
        print('发现项目:' + str(projects.__len__()) + "个")
        print('废弃资源数量:' + str(delList.__len__()))
        for p in projects:
            file = str(p)
            all_path = []
            resList = []
            nums = [0, 0, 0, 0, 0]
            book.save(xlsx)
            #
            init_path(root + '\\' + file)
            for path in all_path:
                scan_file(path)
            sum = 0
            for n in nums:
                sum += n
            nums.insert(0, file)
            nums.append(sum)
            nums.append(resList.__len__())
            sheet.append(nums)
        book.save(xlsx)
    else:
        print('文件:[' + delFile + ']不存在')
    time.sleep(1)
    os.system('pause')
